from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd


def export_xlsx(df: pd.DataFrame, path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise ImportError("openpyxl is required: pip install openpyxl") from e

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Benchmark")
        ws = writer.sheets["Benchmark"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        index_font = Font(bold=True, size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Alternating row shading — group by first index level (model name)
        current_model = None
        shade = False
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            model_cell = row[0].value
            if model_cell is not None:
                current_model = model_cell
                shade = not shade
            for col_idx, cell in enumerate(row):
                if col_idx < 2:
                    cell.font = index_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.number_format = "0.000"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if shade:
                    cell.fill = alt_fill

        for column in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column
            )
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 28)

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "C2"


def export_png(df: pd.DataFrame, path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    display_df = df.reset_index()
    col_labels = list(display_df.columns)
    n_rows = len(display_df)
    n_cols = len(col_labels)
    n_index = len(df.index.names)  # number of index columns (Model, Dataset)

    # Pre-compute per-column min/max for heatmap coloring
    col_ranges: dict = {}
    for col in col_labels[n_index:]:
        vals = pd.to_numeric(display_df[col], errors="coerce").dropna()
        if len(vals) > 1:
            col_ranges[col] = (float(vals.min()), float(vals.max()))

    # Build cell_text and cell_colors using explicit column indexing
    cell_text = []
    cell_colors = []
    for row_i in range(n_rows):
        text_row, color_row = [], []
        for col_i, col in enumerate(col_labels):
            val = display_df.iloc[row_i, col_i]
            if col_i < n_index:
                text_row.append(str(val) if pd.notna(val) else "")
                color_row.append("#EBF3FB")
            else:
                try:
                    fval = float(val)
                    if np.isnan(fval):
                        text_row.append("—")
                        color_row.append("#F5F5F5")
                    else:
                        text_row.append(f"{fval:.3f}")
                        if col in col_ranges:
                            lo, hi = col_ranges[col]
                            norm = (fval - lo) / (hi - lo) if hi > lo else 0.5
                            r = int(255 - norm * 80)
                            g = int(220 + norm * 35)
                            b = int(200 - norm * 60)
                            color_row.append(f"#{r:02X}{g:02X}{b:02X}")
                        else:
                            color_row.append("#FFFFFF")
                except (TypeError, ValueError):
                    text_row.append(str(val) if pd.notna(val) else "—")
                    color_row.append("#FFFFFF")
        assert len(text_row) == n_cols, f"row {row_i}: {len(text_row)} != {n_cols}"
        cell_text.append(text_row)
        cell_colors.append(color_row)

    fig_w = max(10, n_cols * 1.4 + 2)
    fig_h = max(3, n_rows * 0.55 + 1.8)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    table = ax.table(
        cellText=cell_text,
        colLabels=col_labels,
        cellColours=cell_colors,
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.6)

    # Style header row (row 0)
    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor("#1F4E79")
        cell.set_text_props(color="white", fontweight="bold")

    # Bold index columns
    for i in range(n_rows):
        for j in range(n_index):
            table[i + 1, j].set_text_props(fontweight="bold")

    plt.title("Model Benchmark Results", fontsize=13, fontweight="bold", pad=10)
    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close()
